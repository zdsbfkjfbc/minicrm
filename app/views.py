import json
from math import ceil

from flask import (render_template, flash, redirect, url_for, request, Response,
                   abort, jsonify)
from flask_login import current_user, login_user, logout_user, login_required
from app import db
from app.models import User, Contact, AuditLog, SystemSettings
from app.decorators import require_gestor
from app.forms import (ContactForm, DeleteContactForm, ImportForm,
                       LoginForm, LogoutForm, RegistrationForm,
                       SystemSettingsForm)
from app.services.auth import (clear_login_failures, is_login_blocked,
                               register_login_failure, LOGIN_ATTEMPTS)
from app.services.utils import (format_datetime_brt, sanitize_for_spreadsheet,
                                sanitize_html)
from app.tasks import enqueue_import_job, get_job_status
from app.domain.use_cases.calculate_metrics import DashboardMetrics
from app.domain.use_cases.manage_contacts import (CreateContact, DeleteContact,
                                                  GetContact, ListContacts,
                                                  UpdateContact)
from app.domain.use_cases.manage_settings import GetSetting, UpdateSetting
from app.infra.repositories.sqlalchemy_audit_repo import SqlAlchemyAuditRepository
from app.infra.repositories.sqlalchemy_contact_repo import SqlAlchemyContactRepository
from app.infra.repositories.sqlalchemy_settings_repo import SqlAlchemySettingsRepository
from app.infra.repositories.sqlalchemy_user_repo import SqlAlchemyUserRepository
from app.infra.export.pdf_exporter import export_metrics_pdf as generate_metrics_pdf
from flask import current_app as app
from datetime import date, datetime, timezone
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class SimplePagination:
    def __init__(self, items, total: int, page: int, per_page: int):
        self.items = items
        self.total = total
        self.page = page
        self.per_page = per_page

    @property
    def pages(self) -> int:
        return max(1, ceil(self.total / self.per_page)) if self.per_page else 1

    @property
    def has_prev(self) -> bool:
        return self.page > 1

    @property
    def has_next(self) -> bool:
        return self.page < self.pages

    @property
    def prev_num(self) -> int:
        return max(1, self.page - 1)

    @property
    def next_num(self) -> int:
        return min(self.pages, self.page + 1)


def _contact_repo():
    return SqlAlchemyContactRepository()


def _audit_repo():
    return SqlAlchemyAuditRepository()


def _user_repo():
    return SqlAlchemyUserRepository()


def _settings_repo():
    return SqlAlchemySettingsRepository()


def _is_gestor() -> bool:
    return current_user.role == 'Gestor'


def _paginate_contacts(items, total: int, page: int, per_page: int = 15):
    return SimplePagination(items, total, page, per_page)



def record_audit_event(action: str, target_type: str, target_id: int | None = None, details: str | None = None):
    """Registra uma ação no log de auditoria."""
    _audit_repo().save(
        AuditLog(
            user_id=current_user.id,
            action=action,
            target_type=target_type,
            target_id=target_id,
            details=details
        )
    )



def get_authorized_contact(contact_id):
    """Retorna o contato se o usuário tem permissão, senão 404."""
    contact = GetContact(_contact_repo()).execute(
        contact_id, current_user.id, _is_gestor()
    )
    if contact is None:
        abort(404)
    return contact

@app.route('/', methods=['GET', 'POST'])
@app.route('/index', methods=['GET', 'POST'])
@login_required
def index():
    status_filter = request.args.get('status', 'Todos')
    sort_by = request.args.get('sort', 'deadline_asc')
    search_query = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)

    contacts_uc = ListContacts(_contact_repo())
    metrics_uc = DashboardMetrics(_contact_repo(), _user_repo())
    settings_uc = GetSetting(_settings_repo())

    inactive_days = int(settings_uc.execute('days_inactive_alert', '8'))
    items, total = contacts_uc.execute(
        user_id=current_user.id,  
        is_gestor=_is_gestor(),
        status=status_filter,
        search=search_query,
        sort_by=sort_by,
        page=page,
        per_page=15,
    )
    contacts = _paginate_contacts(items, total, page, 15)
    metrics = metrics_uc.index_metrics(
        user_id=current_user.id,
        is_gestor=_is_gestor(),
        inactive_days=inactive_days,
    )
    recent_contacts = _contact_repo().recent(
        user_id=current_user.id,
        is_gestor=_is_gestor(),
        limit=5,
    )

    return render_template('index.html', title='MiniCRM', contacts=contacts, status_filter=status_filter, sort_by=sort_by, search_query=search_query, metrics=metrics, recent_contacts=recent_contacts, inactive_days=inactive_days, delete_form=DeleteContactForm())

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    if form.validate_on_submit():
        blocked, retry_seconds = is_login_blocked(LOGIN_ATTEMPTS, form.username.data)
        if blocked:
            retry_minutes = max(1, retry_seconds // 60)
            flash(f'Muitas tentativas de login. Tente novamente em aproximadamente {retry_minutes} minuto(s).', 'error')
            return redirect(url_for('login'))

        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            register_login_failure(LOGIN_ATTEMPTS, form.username.data)
            app.logger.warning(f"Tentativa de login falha para o usuário: {form.username.data}")
            flash('Usuário ou senha inválidos.')
            return redirect(url_for('login'))
        clear_login_failures(LOGIN_ATTEMPTS, form.username.data)
        login_user(user, remember=form.remember_me.data)
        app.logger.info(f"Login bem-sucedido: {user.username}")
        return redirect(url_for('index'))
    return render_template(
        'login.html',
        title='Entrar',
        form=form,
        hide_nav=True,
        body_class='auth-shell min-h-screen antialiased'
    )

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, role='Operador')
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('Cadastro realizado com sucesso! Agora você pode fazer login.')
        return redirect(url_for('login'))
    return render_template(
        'register.html',
        title='Cadastrar',
        form=form,
        hide_nav=True,
        body_class='auth-shell min-h-screen antialiased'
    )

@app.route('/dashboard')
@require_gestor
def dashboard():
    data = DashboardMetrics(_contact_repo(), _user_repo()).dashboard_full()
    return render_template('dashboard.html', title='Métricas', **data)
@app.route('/logout', methods=['POST'])
@login_required
def logout():
    form = LogoutForm()
    if not form.validate_on_submit():
        flash('Requisição inválida.', 'error')
        return redirect(url_for('index'))
    logout_user()
    return redirect(url_for('login'))

@app.route('/contact/new', methods=['GET', 'POST'])
@login_required
def new_contact():
    form = ContactForm()
    if form.validate_on_submit():
        CreateContact(_contact_repo(), _audit_repo()).execute(
            contact_type=form.contact_type.data,
            customer_name=form.customer_name.data,
            email=form.email.data or None,
            phone=form.phone.data or None,
            status=form.status.data,
            deadline=form.deadline.data,
            observations=sanitize_html(form.observations.data),
            user_id=current_user.id,
        )
        db.session.commit()
        flash('Novo contato adicionado com sucesso!')
        return redirect(url_for('index'))
    return render_template('form.html', title='Novo Contato', form=form, legend='Novo Contato')

@app.route('/contact/<int:contact_id>/view')
@login_required
def view_contact(contact_id):
    """Exibe um contato em modo somente leitura."""
    contact = get_authorized_contact(contact_id)
    return render_template('contact_view.html', title=f'Contato: {contact.customer_name}', contact=contact)

@app.route('/contact/<int:contact_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_contact(contact_id):
    contact = get_authorized_contact(contact_id)
    form = ContactForm()
    if form.validate_on_submit():
        UpdateContact(_contact_repo(), _audit_repo()).execute(
            contact_id=contact_id,
            user_id=current_user.id,
            is_gestor=_is_gestor(),
            contact_type=form.contact_type.data,
            customer_name=form.customer_name.data,
            email=form.email.data or None,
            phone=form.phone.data or None,
            status=form.status.data,
            deadline=form.deadline.data,
            observations=sanitize_html(form.observations.data),
        )
        db.session.commit()
        flash('O contato foi atualizado!')
        return redirect(url_for('index'))
    elif request.method == 'GET':
        form.contact_type.data  = contact.contact_type or 'Pessoa'
        form.customer_name.data = contact.customer_name
        form.email.data         = contact.email
        form.phone.data         = contact.phone
        form.status.data        = contact.status
        form.deadline.data      = contact.deadline
        form.observations.data  = contact.observations
    return render_template('form.html', title='Editar Contato', form=form, legend='Editar Contato')

@app.route('/contact/<int:contact_id>/delete', methods=['POST'])
@login_required
def delete_contact(contact_id):
    form = DeleteContactForm()
    if not form.validate_on_submit():
        abort(400)
    DeleteContact(_contact_repo(), _audit_repo()).execute(
        contact_id, current_user.id, _is_gestor()
    )
    db.session.commit()
    flash('O contato foi excluído.')
    return redirect(url_for('index'))


@app.route('/export/<fmt>')
@login_required
def export_contacts(fmt):
    """Exporta os contatos (respeitando filtro e papel do usuário) em CSV ou XLSX."""
    if fmt not in ('csv', 'xlsx'):
        flash('Formato de exportação inválido.')
        return redirect(url_for('index'))
    status_filter = request.args.get('status', 'Todos')
    search_query = request.args.get('search', '').strip()
    sort_by = request.args.get('sort', 'deadline_asc')
    contacts = _contact_repo().list_all(
        user_id=current_user.id,
        is_gestor=_is_gestor(),
        status=status_filter,
        search=search_query,
        sort_by=sort_by,
    )
    headers = ['ID', 'Tipo', 'Nome / Razão Social', 'E-mail', 'Telefone', 'Status', 'Responsável', 'Prazo', 'Observações', 'Criado em']
    def contact_to_export_row(c):
        return [
            c.id,
            sanitize_for_spreadsheet(c.contact_type or 'Pessoa'),
            sanitize_for_spreadsheet(c.customer_name),
            sanitize_for_spreadsheet(c.email or ''),
            sanitize_for_spreadsheet(c.phone or ''),
            sanitize_for_spreadsheet(c.status),
            sanitize_for_spreadsheet(c.owner_username or ''),
            sanitize_for_spreadsheet(c.deadline.strftime('%d/%m/%Y') if c.deadline else ''),
            sanitize_for_spreadsheet(c.observations or ''),
            sanitize_for_spreadsheet(format_datetime_brt(c.created_at, '%d/%m/%Y %H:%M')),
        ]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        writer.writerow(headers)
        for c in contacts:
            writer.writerow(contact_to_export_row(c))
        output.seek(0)
        return Response(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=atendimentos_{timestamp}.csv'}
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Atendimentos'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='111111')
    header_align = Alignment(horizontal='center', vertical='center')
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    for row_idx, c in enumerate(contacts, start=2):
        for col_idx, value in enumerate(contact_to_export_row(c), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=atendimentos_{timestamp}.xlsx'}
    )


@app.route('/export_metrics/<fmt>')
@login_required
def export_metrics(fmt):
    """Exporta as m?tricas do dashboard em CSV ou XLSX (apenas Gestores)."""
    if current_user.role != 'Gestor':
        flash('Acesso negado.')
        return redirect(url_for('index'))
    if fmt not in ('csv', 'xlsx'):
        flash('Formato inv?lido.')
        return redirect(url_for('dashboard'))
    metrics_data = DashboardMetrics(_contact_repo(), _user_repo()).export_metrics_data()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        writer.writerow(['Resumo de M?tricas', f'Exportado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
        writer.writerow([])
        writer.writerow(['M?trica', 'Valor'])
        writer.writerow(['Total de Atendimentos', metrics_data['total']])
        writer.writerow(['Abertos', metrics_data['abertos']])
        writer.writerow(['Aguardando Cliente', metrics_data['aguardando']])
        writer.writerow(['Resolvidos', metrics_data['resolvidos']])
        writer.writerow(['Cancelados', metrics_data['cancelados']])
        writer.writerow(['Atrasados', metrics_data['overdue']])
        writer.writerow(['SLA (%)', metrics_data['sla_rate']])
        writer.writerow([])
        writer.writerow(['Operador', 'Abertos', 'Aguardando', 'Resolvidos', 'Cancelados'])
        for row in metrics_data['op_rows']:
            writer.writerow([sanitize_for_spreadsheet(row[0]), *row[1:]])
        output.seek(0)
        return Response(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=metricas_{timestamp}.csv'}
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'M?tricas'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='111111')
    summary_data = [
        ('M?trica', 'Valor'),
        ('Total de Atendimentos', metrics_data['total']),
        ('Abertos', metrics_data['abertos']),
        ('Aguardando Cliente', metrics_data['aguardando']),
        ('Resolvidos', metrics_data['resolvidos']),
        ('Cancelados', metrics_data['cancelados']),
        ('Atrasados', metrics_data['overdue']),
        ('SLA (%)', metrics_data['sla_rate']),
    ]
    for r_idx, (label, val) in enumerate(summary_data, start=1):
        c1 = ws.cell(row=r_idx, column=1, value=label)
        c2 = ws.cell(row=r_idx, column=2, value=val)
        if r_idx == 1:
            c1.font = header_font; c1.fill = header_fill
            c2.font = header_font; c2.fill = header_fill
    ws.cell(row=len(summary_data)+2, column=1, value='Por Operador')
    op_headers = ['Operador', 'Abertos', 'Aguardando', 'Resolvidos', 'Cancelados']
    for c_idx, h in enumerate(op_headers, 1):
        cell = ws.cell(row=len(summary_data)+3, column=c_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
    for r_idx, row in enumerate(metrics_data['op_rows'], start=len(summary_data)+4):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=metricas_{timestamp}.xlsx'}
    )


@app.route('/export_metrics/pdf')
@require_gestor
def export_metrics_pdf():
    """Exporta as métricas do dashboard em PDF (apenas Gestores)."""
    metrics_data = DashboardMetrics(_contact_repo(), _user_repo()).export_metrics_data()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    pdf_bytes = generate_metrics_pdf(metrics_data)
    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=metricas_{timestamp}.pdf'}
    )


@app.route('/export_bi')
@require_gestor
def export_bi():
    """Exporta todos os contatos em CSV plano para Power BI / Looker Studio (apenas Gestores)."""
    contacts = _contact_repo().list_all(
        user_id=current_user.id, is_gestor=True, sort_by='created_desc'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')

    BI_FLAT_HEADERS = [
        'id', 'tipo', 'nome', 'email', 'telefone',
        'status', 'responsavel', 'prazo', 'observacoes',
        'criado_em', 'atrasado'
    ]
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(BI_FLAT_HEADERS)
    for c in contacts:
        writer.writerow([
            c.id,
            sanitize_for_spreadsheet(c.contact_type or 'Pessoa'),
            sanitize_for_spreadsheet(c.customer_name),
            sanitize_for_spreadsheet(c.email or ''),
            sanitize_for_spreadsheet(c.phone or ''),
            sanitize_for_spreadsheet(c.status),
            sanitize_for_spreadsheet(c.owner_username or ''),
            sanitize_for_spreadsheet(c.deadline.strftime('%Y-%m-%d') if c.deadline else ''),
            sanitize_for_spreadsheet(c.observations or ''),
            sanitize_for_spreadsheet(format_datetime_brt(c.created_at, '%Y-%m-%d %H:%M:%S')),
            '1' if c.is_overdue() else '0',
        ])
    output.seek(0)
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=minicrm_bi_{timestamp}.csv'}
    )


@app.route('/audit_logs')
@require_gestor
def audit_logs():
    """Página de logs de auditoria (apenas Gestores)."""
    page = request.args.get('page', 1, type=int)
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).paginate(page=page, per_page=30, error_out=False)
    return render_template('audit_logs.html', title='Log de Auditoria', logs=logs)
@app.route('/webhooks/notify', methods=['POST'])
def webhook_notify():
    token = request.headers.get('X-Webhook-Token', '')
    if token != app.config['WEBHOOK_TOKEN']:
        abort(403)
    payload = request.get_json(silent=True) or {}
    details = json.dumps(payload, default=str)
    record_audit_event('webhook', 'Webhook', None, details=details)
    app.logger.info(f"Webhook recebido (%s)", payload)
    return jsonify({'status': 'accepted'}), 202


@app.route('/alerts')
@require_gestor
def alert_history():
    alerts = AuditLog.query.filter_by(action='webhook').order_by(AuditLog.timestamp.desc()).limit(50).all()
    return render_template('alerts.html', alerts=alerts)


@app.route('/system_settings', methods=['GET', 'POST'])
@require_gestor
def system_settings():
    """Configurações do sistema (apenas Gestores)."""
    settings_uc = GetSetting(_settings_repo())
    update_uc = UpdateSetting(_settings_repo(), _audit_repo())

    form = SystemSettingsForm()
    days_inactive = settings_uc.execute('days_inactive_alert', '8')
    if request.method == 'GET':
        try:
            form.days_inactive_alert.data = int(days_inactive)
        except ValueError:
            form.days_inactive_alert.data = 8

    if form.validate_on_submit():
        days_value = str(form.days_inactive_alert.data)
        update_uc.execute(
            'days_inactive_alert', days_value,
            user_id=current_user.id,
            description='Número de dias sem atividade para gerar alerta de inatividade'
        )
        db.session.commit()
        flash('Configurações salvas com sucesso!')
        return redirect(url_for('system_settings'))

    return render_template('system_settings.html', title='Configurações', days_inactive=days_inactive, form=form)


@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_csv():
    form = ImportForm()
    if form.validate_on_submit():
        file = form.csv_file.data
        content = file.stream.read().decode("UTF8")
        job_id = enqueue_import_job(content, current_user.id)
        flash(f'Importação agendada (job {job_id}).', 'info')
        return redirect(url_for('import_status', job_id=job_id))
    return render_template('import.html', title='Importar Clientes', form=form)


@app.route('/import/status/<job_id>')
@login_required
def import_status(job_id):
    status = get_job_status(job_id, current_user.id)
    if status is None:
        abort(404)
    return render_template('import_status.html', job_id=job_id, status=status)



