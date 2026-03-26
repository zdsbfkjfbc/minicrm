"""XLSX export functions for contacts and metrics."""

from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.domain.entities.contact import Contact
from app.services.utils import format_datetime_brt, sanitize_for_spreadsheet

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="111111")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

CONTACT_HEADERS = [
    "ID", "Tipo", "Nome / Razão Social", "E-mail", "Telefone",
    "Status", "Responsável", "Prazo", "Observações", "Criado em",
]


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _contact_row(c: Contact) -> list:
    return [
        c.id,
        sanitize_for_spreadsheet(c.contact_type or "Pessoa"),
        sanitize_for_spreadsheet(c.customer_name),
        sanitize_for_spreadsheet(c.email or ""),
        sanitize_for_spreadsheet(c.phone or ""),
        sanitize_for_spreadsheet(c.status),
        sanitize_for_spreadsheet(c.owner_username or ""),
        sanitize_for_spreadsheet(
            c.deadline.strftime("%d/%m/%Y") if c.deadline else ""
        ),
        sanitize_for_spreadsheet(c.observations or ""),
        sanitize_for_spreadsheet(
            format_datetime_brt(c.created_at, "%d/%m/%Y %H:%M")
        ),
    ]


def export_contacts_xlsx(contacts: list[Contact]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atendimentos"

    for col_idx, h in enumerate(CONTACT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, c in enumerate(contacts, start=2):
        for col_idx, value in enumerate(_contact_row(c), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _auto_width(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_metrics_xlsx(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Métricas"

    summary_data = [
        ("Métrica", "Valor"),
        ("Total de Atendimentos", data["total"]),
        ("Abertos", data["abertos"]),
        ("Aguardando Cliente", data["aguardando"]),
        ("Resolvidos", data["resolvidos"]),
        ("Cancelados", data["cancelados"]),
        ("Atrasados", data["overdue"]),
        ("SLA (%)", data["sla_rate"]),
    ]
    for r_idx, (label, val) in enumerate(summary_data, start=1):
        c1 = ws.cell(row=r_idx, column=1, value=label)
        c2 = ws.cell(row=r_idx, column=2, value=val)
        if r_idx == 1:
            c1.font = HEADER_FONT
            c1.fill = HEADER_FILL
            c2.font = HEADER_FONT
            c2.fill = HEADER_FILL

    row_offset = len(summary_data) + 2
    ws.cell(row=row_offset, column=1, value="Por Operador")
    op_headers = [
        "Operador", "Abertos", "Aguardando", "Resolvidos", "Cancelados"
    ]
    for c_idx, h in enumerate(op_headers, 1):
        cell = ws.cell(row=row_offset + 1, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for r_idx, row in enumerate(data["op_rows"], start=row_offset + 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    _auto_width(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
